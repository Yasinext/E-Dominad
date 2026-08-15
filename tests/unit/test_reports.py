from __future__ import annotations

from datetime import UTC, datetime

from openpyxl import load_workbook

from domainbot.domain.parser import ReportFilter
from domainbot.reports.excel import (
    safe_general_report_filename,
    safe_report_filename,
    write_excel_report,
)
from domainbot.reports.text import render_expiration_report, render_text_report
from domainbot.reports.types import Report, ReportRow


def _report() -> Report:
    checked_at = datetime(2026, 7, 30, 16, 0, tzinfo=UTC)
    return Report(
        chat_id=-1,
        root="test",
        range_start=1,
        range_end=3,
        job_id="job-1",
        finished_at=checked_at,
        report_filter=ReportFilter.ALL,
        rows=(
            ReportRow(
                ordinal=1,
                domain="test1.com",
                verified_status="REGISTERED",
                last_check_outcome="REGISTERED",
                registration_date=None,
                expiration_date=datetime(2027, 2, 18, 10, 12, 35, tzinfo=UTC),
                registrar_name="Example Registrar",
                registrar_iana_id="999",
                rdap_statuses=("client transfer prohibited",),
                nameservers=("ns1.example.com",),
                http_status=200,
                attempt_count=1,
                response_time_ms=210,
                checked_at=checked_at,
                btk_status="clear",
                btk_checked_at=checked_at,
                btk_note="TR DNS gercek IP donuyor, sinkhole yok.",
                btk_error=None,
                explanation="RDAP kaydı bulundu.",
            ),
            ReportRow(
                ordinal=2,
                domain="test2.com",
                verified_status="NOT_FOUND_IN_REGISTRY",
                last_check_outcome="NOT_FOUND_IN_REGISTRY",
                registration_date=None,
                expiration_date=None,
                registrar_name=None,
                registrar_iana_id=None,
                rdap_statuses=(),
                nameservers=(),
                http_status=404,
                attempt_count=1,
                response_time_ms=180,
                checked_at=checked_at,
                btk_status=None,
                btk_checked_at=None,
                btk_note=None,
                btk_error=None,
                explanation="Registry kaydı bulunamadı.",
            ),
            ReportRow(
                ordinal=3,
                domain="test3.com",
                verified_status="REGISTERED",
                last_check_outcome="REGISTERED",
                registration_date=None,
                expiration_date=None,
                registrar_name=None,
                registrar_iana_id=None,
                rdap_statuses=(),
                nameservers=(),
                http_status=200,
                attempt_count=1,
                response_time_ms=180,
                checked_at=checked_at,
                btk_status="inconclusive",
                btk_checked_at=checked_at,
                btk_note="TR DNS cozulemedi, sonuc belirsiz.",
                btk_error=None,
                explanation="RDAP kaydı bulundu.",
            ),
        ),
    )


def test_render_text_report_includes_summary_and_rows() -> None:
    text = render_text_report(_report(), row_limit=20)

    assert "Rapor hazır." in text
    assert "Kök: test" in text
    assert "Kayıtlı: 2" in text
    assert "Registry kaydı bulunamadı: 1" in text
    assert "Son sorgu: 30.07.2026 19:00:00" in text
    assert "1. test1.com - Alınmış: Evet - Son geçerlilik: 18.02.2027 - DNS: Var" in text


def test_render_general_text_report() -> None:
    source = _report()
    report = Report(
        chat_id=source.chat_id,
        root=None,
        range_start=None,
        range_end=None,
        job_id="general",
        finished_at=None,
        report_filter=ReportFilter.ALL,
        rows=source.rows,
    )

    text = render_text_report(report, row_limit=20)

    assert "Genel rapor hazır." in text
    assert "Kök:" not in text
    assert "Havuzdaki Domain Sayısı: 3" in text
    assert "BTK kontrollü: 2" in text
    assert "BTK kontrol bekleyen:" not in text
    assert "Kaynak: RDAP" not in text


def test_render_general_text_report_uses_excel_hint_for_large_reports() -> None:
    source = _report()
    report = Report(
        chat_id=source.chat_id,
        root=None,
        range_start=None,
        range_end=None,
        job_id="general",
        finished_at=None,
        report_filter=ReportFilter.ALL,
        rows=source.rows,
    )

    text = render_text_report(report, row_limit=1)

    assert "Detaylı rapor için excel kullanın." in text


def test_render_expiration_report_includes_ordered_rows() -> None:
    text = render_expiration_report(_report(), row_limit=20)

    assert "Geçerlilik raporu hazır." in text
    assert "1. test1.com - Kayıtlı - Bitiş: 18.02.2027" in text
    assert "2. test2.com - Registry kaydı bulunamadı - Bitiş: -" in text


def test_write_excel_report_creates_expected_sheets(tmp_path) -> None:  # type: ignore[no-untyped-def]
    path = tmp_path / "report.xlsx"

    write_excel_report(_report(), path)

    workbook = load_workbook(path)
    assert workbook.sheetnames == ["Genel Rapor"]
    assert workbook["Genel Rapor"]["B2"].value == "test1.com"
    assert workbook["Genel Rapor"]["C2"].value == "Evet"
    assert workbook["Genel Rapor"]["E2"].value == "18.02.2027"
    assert workbook["Genel Rapor"]["F2"].value == "Evet"
    assert workbook["Genel Rapor"]["K2"].value == "Engelsiz"
    assert workbook["Genel Rapor"]["L2"].value == (
        "Domain kayıtlı; registry tarafında nameserver bilgisi var."
    )
    assert workbook["Genel Rapor"]["K3"].value == "Domain kayıtlı değil"
    assert workbook["Genel Rapor"]["K4"].value == "Engelsiz"


def test_safe_report_filename_sanitizes_root() -> None:
    assert safe_report_filename("test-1", 1, 5, "20260730T160000Z") == (
        "domain-report_test-1_1-5_20260730T160000Z.xlsx"
    )


def test_safe_general_report_filename() -> None:
    assert safe_general_report_filename("20260730T160000Z") == (
        "domain-report_genel_20260730T160000Z.xlsx"
    )
